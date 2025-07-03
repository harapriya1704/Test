from openpyxl import load_workbook
from config import EXCEL_PATH

def extract_glassbox_data(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    headers = {str(cell.value).strip().replace(" ", "_").lower(): idx for idx, cell in enumerate(ws[1])}

    col_fiscal_week = headers.get("fiscal_week")
    col_date = headers.get("date")
    col_order_number = headers.get("order_number")
    col_improve_text = headers.get("improve_text")
    col_glassbox_link = headers.get("glassbox_link")
    col_sat_dissat = headers.get("sat/dissat")

    required_columns = {
        "Fiscal Week": col_fiscal_week,
        "Date": col_date,
        "Order Number": col_order_number,
        "Improve Text": col_improve_text,
        "Glassbox Link": col_glassbox_link,
        "Sat/Dissat": col_sat_dissat
    }

    for name, idx in required_columns.items():
        if idx is None:
            raise ValueError(f"Missing column: {name}")

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_number = row[col_order_number]
        sat_dissat = row[col_sat_dissat]
        if order_number and str(sat_dissat).strip().upper() == 'DSAT':
            data.append({
                "fiscal_week": row[col_fiscal_week],
                "date": row[col_date],
                "order_number": order_number,
                "improve_text": row[col_improve_text],
                "glassbox_link": row[col_glassbox_link],
                "sat_dissat": sat_dissat
            })

    return data
